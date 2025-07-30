from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, flash, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from datetime import datetime, timedelta, date
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.security import generate_password_hash, check_password_hash
from config import Config
from notifications import mail, send_notification_retard, send_notification_absence
import plotly.utils
import pyotp
import qrcode
import io
import base64
from pdf_generator import generate_payslip

app = Flask(__name__)
app.config.from_object(Config)

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Initialiser Flask-Mail
mail.init_app(app)

# Note: On utilise l'heure locale du système sans timezone pour éviter les problèmes avec SQLite

# Modèles de base de données
class Admin(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    totp_secret = db.Column(db.String(32))
    is_2fa_enabled = db.Column(db.Boolean, default=False)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    def generate_totp_secret(self):
        """Génère un nouveau secret TOTP"""
        self.totp_secret = pyotp.random_base32()
        return self.totp_secret
    
    def get_totp_uri(self):
        """Génère l'URI pour le QR code"""
        return pyotp.totp.TOTP(self.totp_secret).provisioning_uri(
            name=self.username,
            issuer_name='Globibat Badge System'
        )
    
    def verify_totp(self, token):
        """Vérifie le code TOTP"""
        if not self.totp_secret:
            return False
        totp = pyotp.TOTP(self.totp_secret)
        return totp.verify(token, valid_window=1)

class Employe(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    matricule = db.Column(db.String(50), unique=True, nullable=False)
    nom = db.Column(db.String(100), nullable=False)
    prenom = db.Column(db.String(100), nullable=False)
    departement = db.Column(db.String(100))
    email = db.Column(db.String(120))
    telephone = db.Column(db.String(20))
    date_embauche = db.Column(db.Date, default=date.today)
    actif = db.Column(db.Boolean, default=True)
    pointages = db.relationship('Pointage', backref='employe', lazy=True)

class Pointage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employe_id = db.Column(db.Integer, db.ForeignKey('employe.id'), nullable=False)
    date_pointage = db.Column(db.Date, nullable=False, default=date.today)
    
    # Les 4 moments de badgeage
    arrivee_matin = db.Column(db.DateTime)
    depart_midi = db.Column(db.DateTime)
    arrivee_apres_midi = db.Column(db.DateTime)
    depart_soir = db.Column(db.DateTime)
    
    # Informations supplémentaires
    heures_travaillees = db.Column(db.Float, default=0)
    retard_matin = db.Column(db.Boolean, default=False)
    retard_apres_midi = db.Column(db.Boolean, default=False)
    absence = db.Column(db.Boolean, default=False)
    commentaire = db.Column(db.Text)
    
    # Contrainte d'unicité : un seul enregistrement par employé par jour
    __table_args__ = (db.UniqueConstraint('employe_id', 'date_pointage'),)

@login_manager.user_loader
def load_user(user_id):
    return Admin.query.get(int(user_id))

# Fonction pour calculer les heures travaillées
def calculer_heures(pointage):
    heures_total = 0
    
    # Période du matin
    if pointage.arrivee_matin and pointage.depart_midi:
        delta_matin = pointage.depart_midi - pointage.arrivee_matin
        heures_total += delta_matin.total_seconds() / 3600
    
    # Période de l'après-midi
    if pointage.arrivee_apres_midi and pointage.depart_soir:
        delta_apres_midi = pointage.depart_soir - pointage.arrivee_apres_midi
        heures_total += delta_apres_midi.total_seconds() / 3600
    
    return round(heures_total, 2)

# Routes principales
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/badge', methods=['GET', 'POST'])
def badge():
    if request.method == 'POST':
        matricule = request.form.get('matricule')
        type_badge = request.form.get('type_badge')
        
        employe = Employe.query.filter_by(matricule=matricule, actif=True).first()
        
        if not employe:
            flash('Matricule invalide ou employé inactif', 'error')
            return redirect(url_for('badge'))
        
        # Obtenir ou créer le pointage du jour
        aujourd_hui = date.today()
        pointage = Pointage.query.filter_by(
            employe_id=employe.id,
            date_pointage=aujourd_hui
        ).first()
        
        if not pointage:
            pointage = Pointage(
                employe_id=employe.id,
                date_pointage=aujourd_hui
            )
            db.session.add(pointage)
        
        # Enregistrer l'heure actuelle (sans timezone pour SQLite)
        maintenant = datetime.now()
        
        # Logique de badgeage selon le type
        if type_badge == 'arrivee_matin':
            if pointage.arrivee_matin:
                flash('Vous avez déjà badgé votre arrivée ce matin', 'warning')
            else:
                pointage.arrivee_matin = maintenant
                # Vérifier le retard (exemple : après 9h00)
                heure_limite = datetime.strptime(app.config.get('HEURE_ARRIVEE_MAX', '09:00'), '%H:%M').time()
                if maintenant.time() > heure_limite:
                    pointage.retard_matin = True
                    # Envoyer notification de retard
                    heure_attendue = datetime.combine(date.today(), heure_limite)
                    send_notification_retard(employe, maintenant, heure_attendue)
                flash(f'Arrivée enregistrée à {maintenant.strftime("%H:%M")}', 'success')
        
        elif type_badge == 'depart_midi':
            if not pointage.arrivee_matin:
                flash('Vous devez d\'abord badger votre arrivée', 'error')
            elif pointage.depart_midi:
                flash('Vous avez déjà badgé votre départ midi', 'warning')
            else:
                pointage.depart_midi = maintenant
                flash(f'Départ midi enregistré à {maintenant.strftime("%H:%M")}', 'success')
        
        elif type_badge == 'arrivee_apres_midi':
            if pointage.arrivee_apres_midi:
                flash('Vous avez déjà badgé votre retour de pause', 'warning')
            else:
                pointage.arrivee_apres_midi = maintenant
                # Vérifier le retard (exemple : après 14h00)
                heure_limite = datetime.strptime(app.config.get('HEURE_RETOUR_MAX', '14:00'), '%H:%M').time()
                if maintenant.time() > heure_limite:
                    pointage.retard_apres_midi = True
                    # Envoyer notification de retard
                    heure_attendue = datetime.combine(date.today(), heure_limite)
                    send_notification_retard(employe, maintenant, heure_attendue)
                flash(f'Retour de pause enregistré à {maintenant.strftime("%H:%M")}', 'success')
        
        elif type_badge == 'depart_soir':
            if not pointage.arrivee_matin and not pointage.arrivee_apres_midi:
                flash('Vous devez d\'abord badger une arrivée', 'error')
            elif pointage.depart_soir:
                flash('Vous avez déjà badgé votre départ', 'warning')
            else:
                pointage.depart_soir = maintenant
                # Calculer les heures travaillées
                pointage.heures_travaillees = calculer_heures(pointage)
                flash(f'Départ enregistré à {maintenant.strftime("%H:%M")}. Total: {pointage.heures_travaillees}h', 'success')
        
        db.session.commit()
        return redirect(url_for('badge'))
    
    return render_template('badge.html')

# NOUVEAU : Portail Employé
@app.route('/employe')
def portail_employe():
    """Page d'accueil du portail employé"""
    return render_template('employe_home.html')

@app.route('/employe/login', methods=['GET', 'POST'])
def employe_login():
    """Connexion employé avec matricule"""
    if request.method == 'POST':
        matricule = request.form.get('matricule')
        employe = Employe.query.filter_by(matricule=matricule, actif=True).first()
        
        if employe:
            # Stocker l'ID employé en session
            session['employe_id'] = employe.id
            session['employe_matricule'] = employe.matricule
            return redirect(url_for('employe_dashboard'))
        else:
            flash('Matricule invalide ou employé inactif', 'error')
    
    return render_template('employe_login.html')

@app.route('/employe/dashboard')
def employe_dashboard():
    """Tableau de bord employé"""
    if 'employe_id' not in session:
        return redirect(url_for('employe_login'))
    
    employe = Employe.query.get(session['employe_id'])
    if not employe:
        session.clear()
        return redirect(url_for('employe_login'))
    
    # Récupérer les pointages du mois en cours
    debut_mois = date.today().replace(day=1)
    pointages = Pointage.query.filter_by(employe_id=employe.id).filter(
        Pointage.date_pointage >= debut_mois
    ).order_by(Pointage.date_pointage.desc()).all()
    
    # Calculer les statistiques
    total_heures = sum(p.heures_travaillees for p in pointages)
    nb_retards = sum(1 for p in pointages if p.retard_matin or p.retard_apres_midi)
    
    return render_template('employe_dashboard.html',
                         employe=employe,
                         pointages=pointages,
                         total_heures=total_heures,
                         nb_retards=nb_retards)

@app.route('/employe/logout')
def employe_logout():
    """Déconnexion employé"""
    session.pop('employe_id', None)
    session.pop('employe_matricule', None)
    return redirect(url_for('portail_employe'))

@app.route('/admin')
@login_required
def admin_dashboard():
    # Statistiques générales
    nb_employes = Employe.query.filter_by(actif=True).count()
    
    # Pointages du jour
    aujourd_hui = date.today()
    pointages_jour = db.session.query(Pointage, Employe).join(Employe).filter(
        Pointage.date_pointage == aujourd_hui
    ).all()
    
    # Calcul des présents/absents
    employes_actifs = Employe.query.filter_by(actif=True).all()
    employes_badges = [p.employe_id for p, e in pointages_jour]
    absents = [e for e in employes_actifs if e.id not in employes_badges]
    
    return render_template('admin_dashboard.html',
                         nb_employes=nb_employes,
                         pointages_jour=pointages_jour,
                         absents=absents,
                         aujourd_hui=aujourd_hui)

@app.route('/admin/employes')
@login_required
def liste_employes():
    employes = Employe.query.order_by(Employe.nom).all()
    return render_template('liste_employes.html', employes=employes)

@app.route('/admin/employe/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter_employe():
    if request.method == 'POST':
        employe = Employe(
            matricule=request.form.get('matricule'),
            nom=request.form.get('nom'),
            prenom=request.form.get('prenom'),
            departement=request.form.get('departement'),
            email=request.form.get('email'),
            telephone=request.form.get('telephone')
        )
        
        try:
            db.session.add(employe)
            db.session.commit()
            flash('Employé ajouté avec succès', 'success')
            return redirect(url_for('liste_employes'))
        except:
            db.session.rollback()
            flash('Erreur: Le matricule existe déjà', 'error')
    
    return render_template('ajouter_employe.html')

@app.route('/admin/rapport/<type_rapport>')
@login_required
def generer_rapport(type_rapport):
    if type_rapport == 'jour':
        date_str = request.args.get('date', date.today().isoformat())
        date_rapport = datetime.fromisoformat(date_str).date()
        
        pointages = db.session.query(Pointage, Employe).join(Employe).filter(
            Pointage.date_pointage == date_rapport
        ).all()
        
        # Créer le fichier Excel avec openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Pointages"
        
        # En-têtes
        headers = ['Matricule', 'Nom', 'Prénom', 'Département', 'Arrivée Matin', 
                   'Départ Midi', 'Arrivée Après-midi', 'Départ Soir', 
                   'Heures Travaillées', 'Retard Matin', 'Retard Après-midi']
        
        # Style pour les en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajouter les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Ajouter les données
        row_num = 2
        for p, e in pointages:
            ws.cell(row=row_num, column=1, value=e.matricule)
            ws.cell(row=row_num, column=2, value=e.nom)
            ws.cell(row=row_num, column=3, value=e.prenom)
            ws.cell(row=row_num, column=4, value=e.departement or '-')
            ws.cell(row=row_num, column=5, value=p.arrivee_matin.strftime('%H:%M') if p.arrivee_matin else 'Non badgé')
            ws.cell(row=row_num, column=6, value=p.depart_midi.strftime('%H:%M') if p.depart_midi else 'Non badgé')
            ws.cell(row=row_num, column=7, value=p.arrivee_apres_midi.strftime('%H:%M') if p.arrivee_apres_midi else 'Non badgé')
            ws.cell(row=row_num, column=8, value=p.depart_soir.strftime('%H:%M') if p.depart_soir else 'Non badgé')
            ws.cell(row=row_num, column=9, value=p.heures_travaillees)
            ws.cell(row=row_num, column=10, value='Oui' if p.retard_matin else 'Non')
            ws.cell(row=row_num, column=11, value='Oui' if p.retard_apres_midi else 'Non')
            row_num += 1
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder le fichier
        filename = f'rapport_pointage_{date_rapport.isoformat()}.xlsx'
        filepath = os.path.join('static', filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True)
    
    elif type_rapport == 'mensuel':
        # NOUVEAU : Rapport mensuel
        mois = int(request.args.get('mois', date.today().month))
        annee = int(request.args.get('annee', date.today().year))
        
        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Rapport {mois:02d}-{annee}"
        
        # Titre
        ws.merge_cells('A1:H1')
        titre = ws['A1']
        titre.value = f"RAPPORT MENSUEL DE PRÉSENCE - {mois:02d}/{annee}"
        titre.font = Font(size=16, bold=True)
        titre.alignment = Alignment(horizontal="center")
        
        # Récupérer tous les employés actifs
        employes = Employe.query.filter_by(actif=True).order_by(Employe.nom).all()
        
        row_num = 3
        for employe in employes:
            # Section employé
            ws.merge_cells(f'A{row_num}:H{row_num}')
            cell_employe = ws[f'A{row_num}']
            cell_employe.value = f"{employe.nom} {employe.prenom} - Matricule: {employe.matricule}"
            cell_employe.font = Font(bold=True, size=12)
            cell_employe.fill = PatternFill("solid", fgColor="E0E0E0")
            row_num += 1
            
            # Statistiques du mois
            pointages = Pointage.query.filter_by(employe_id=employe.id).filter(
                db.extract('month', Pointage.date_pointage) == mois,
                db.extract('year', Pointage.date_pointage) == annee
            ).all()
            
            total_heures = sum(p.heures_travaillees for p in pointages)
            nb_retards = sum(1 for p in pointages if p.retard_matin or p.retard_apres_midi)
            nb_absences = sum(1 for p in pointages if p.absence)
            
            # Ligne de stats
            ws[f'A{row_num}'] = "Total heures:"
            ws[f'B{row_num}'] = f"{total_heures:.2f}h"
            ws[f'C{row_num}'] = "Retards:"
            ws[f'D{row_num}'] = nb_retards
            ws[f'E{row_num}'] = "Absences:"
            ws[f'F{row_num}'] = nb_absences
            row_num += 2
        
        # Sauvegarder
        filename = f'rapport_mensuel_{annee}_{mois:02d}.xlsx'
        filepath = os.path.join('static', filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True)
    
    return redirect(url_for('admin_dashboard'))

# NOUVEAU : API pour les statistiques
@app.route('/api/stats/dashboard')
@login_required
def api_stats_dashboard():
    """API pour obtenir les statistiques du tableau de bord"""
    # Stats du jour
    aujourd_hui = date.today()
    pointages_jour = Pointage.query.filter_by(date_pointage=aujourd_hui).count()
    
    # Stats de la semaine
    debut_semaine = aujourd_hui - timedelta(days=aujourd_hui.weekday())
    pointages_semaine = Pointage.query.filter(
        Pointage.date_pointage >= debut_semaine
    ).all()
    
    heures_semaine = sum(p.heures_travaillees for p in pointages_semaine)
    
    # Stats du mois
    debut_mois = aujourd_hui.replace(day=1)
    pointages_mois = Pointage.query.filter(
        Pointage.date_pointage >= debut_mois
    ).all()
    
    heures_mois = sum(p.heures_travaillees for p in pointages_mois)
    retards_mois = sum(1 for p in pointages_mois if p.retard_matin or p.retard_apres_midi)
    
    return jsonify({
        'jour': {
            'presents': pointages_jour,
            'date': aujourd_hui.strftime('%d/%m/%Y')
        },
        'semaine': {
            'heures_totales': round(heures_semaine, 2),
            'debut': debut_semaine.strftime('%d/%m/%Y')
        },
        'mois': {
            'heures_totales': round(heures_mois, 2),
            'retards': retards_mois,
            'mois': aujourd_hui.strftime('%B %Y')
        }
    })

# NOUVEAU : Page de statistiques avancées
@app.route('/admin/statistiques')
@login_required
def statistiques_avancees():
    """Page de statistiques avancées avec graphiques"""
    from charts import create_dashboard_charts
    
    # Générer les graphiques
    charts = create_dashboard_charts(db.session)
    
    # Statistiques supplémentaires
    employes_actifs = Employe.query.filter_by(actif=True).count()
    
    # Top 5 des employés les plus ponctuels ce mois
    debut_mois = date.today().replace(day=1)
    top_ponctuels = db.session.query(
        Employe,
        db.func.count(Pointage.id).label('jours_presents')
    ).join(Pointage).filter(
        Pointage.date_pointage >= debut_mois,
        Pointage.retard_matin == False,
        Pointage.retard_apres_midi == False
    ).group_by(Employe.id).order_by(
        db.desc('jours_presents')
    ).limit(5).all()
    
    return render_template('statistiques.html',
                         charts=charts,
                         employes_actifs=employes_actifs,
                         top_ponctuels=top_ponctuels)

@app.route('/login', methods=['GET', 'POST'])
@app.route('/admin-globibat', methods=['GET', 'POST'])  # URL secrète
def login():
    if current_user.is_authenticated:
        return redirect(url_for('admin_dashboard'))
        
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        admin = Admin.query.filter_by(username=username).first()
        
        if admin and admin.check_password(password):
            # Si 2FA est activé, rediriger vers la page 2FA
            if admin.is_2fa_enabled:
                session['pending_admin_id'] = admin.id
                return redirect(url_for('verify_2fa'))
            else:
                login_user(admin, remember=True)
                next_page = request.args.get('next')
                return redirect(next_page) if next_page else redirect(url_for('admin_dashboard'))
        else:
            flash('Identifiants invalides', 'error')
    
    return render_template('login.html')

@app.route('/verify-2fa', methods=['GET', 'POST'])
def verify_2fa():
    """Page de vérification 2FA"""
    if 'pending_admin_id' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        code = request.form.get('code')
        admin = Admin.query.get(session['pending_admin_id'])
        
        if admin and admin.verify_totp(code):
            session.pop('pending_admin_id', None)
            login_user(admin, remember=True)
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Code invalide ou expiré', 'error')
    
    return render_template('verify_2fa.html')

@app.route('/admin/setup-2fa')
@login_required
def setup_2fa():
    """Page de configuration 2FA"""
    if current_user.is_2fa_enabled:
        flash('2FA déjà activé', 'info')
        return redirect(url_for('admin_dashboard'))
    
    # Générer un nouveau secret
    secret = current_user.generate_totp_secret()
    db.session.commit()
    
    # Générer le QR code
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(current_user.get_totp_uri())
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    
    qr_code = base64.b64encode(buf.getvalue()).decode()
    
    return render_template('setup_2fa.html', 
                         secret=secret,
                         qr_code=qr_code)

@app.route('/admin/confirm-2fa', methods=['POST'])
@login_required
def confirm_2fa():
    """Confirme l'activation du 2FA"""
    code = request.form.get('code')
    
    if current_user.verify_totp(code):
        current_user.is_2fa_enabled = True
        db.session.commit()
        flash('2FA activé avec succès !', 'success')
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': 'Code invalide'})

@app.route('/admin/disable-2fa', methods=['POST'])
@login_required
def disable_2fa():
    """Désactive le 2FA"""
    password = request.form.get('password')
    
    if current_user.check_password(password):
        current_user.is_2fa_enabled = False
        current_user.totp_secret = None
        db.session.commit()
        flash('2FA désactivé', 'info')
        return redirect(url_for('admin_dashboard'))
    else:
        flash('Mot de passe incorrect', 'error')
        return redirect(url_for('admin_dashboard'))

# NOUVEAU : Génération de fiches de paie
@app.route('/admin/fiche-paie/<int:employe_id>')
@login_required
def generer_fiche_paie(employe_id):
    """Génère une fiche de paie PDF pour un employé"""
    mois = request.args.get('mois', date.today().month, type=int)
    annee = request.args.get('annee', date.today().year, type=int)
    
    # Récupérer l'employé
    employe = Employe.query.get_or_404(employe_id)
    
    # Calculer les heures travaillées ce mois
    pointages = Pointage.query.filter_by(employe_id=employe.id).filter(
        db.extract('month', Pointage.date_pointage) == mois,
        db.extract('year', Pointage.date_pointage) == annee
    ).all()
    
    heures_travaillees = sum(p.heures_travaillees for p in pointages)
    
    # Générer le PDF
    pdf_buffer = generate_payslip(employe, mois, annee, heures_travaillees)
    
    # Nom du fichier
    filename = f'fiche_paie_{employe.matricule}_{annee}_{mois:02d}.pdf'
    
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@app.route('/admin/fiches-paie')
@login_required
def liste_fiches_paie():
    """Page pour générer les fiches de paie"""
    employes = Employe.query.filter_by(actif=True).order_by(Employe.nom).all()
    
    # Mois et année actuels
    mois_actuel = date.today().month
    annee_actuelle = date.today().year
    
    # Calculer les heures pour chaque employé ce mois
    employes_data = []
    for emp in employes:
        pointages = Pointage.query.filter_by(employe_id=emp.id).filter(
            db.extract('month', Pointage.date_pointage) == mois_actuel,
            db.extract('year', Pointage.date_pointage) == annee_actuelle
        ).all()
        
        heures = sum(p.heures_travaillees for p in pointages)
        employes_data.append({
            'employe': emp,
            'heures': heures,
            'salaire_net': heures * 15.0 * 0.77  # Approximation
        })
    
    return render_template('fiches_paie.html',
                         employes_data=employes_data,
                         mois=mois_actuel,
                         annee=annee_actuelle)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

# Initialisation de la base de données
def init_db():
    with app.app_context():
        db.create_all()
        
        # Ne pas créer d'admin automatiquement

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000) 