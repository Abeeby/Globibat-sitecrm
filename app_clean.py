from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, flash, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from datetime import datetime, timedelta, date
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.security import generate_password_hash, check_password_hash
from config import Config

app = Flask(__name__)
app.config.from_object(Config)

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Modèles de base de données
class Admin(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Employe(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    matricule = db.Column(db.String(50), unique=True, nullable=False)
    nom = db.Column(db.String(100), nullable=False)
    prenom = db.Column(db.String(100), nullable=False)
    departement = db.Column(db.String(100))
    email = db.Column(db.String(120))
    telephone = db.Column(db.String(20))
    date_embauche = db.Column(db.Date, default=date.today)
    actif = db.Column(db.Boolean, default=True)
    pointages = db.relationship('Pointage', backref='employe', lazy=True, cascade='all, delete-orphan')

class Pointage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employe_id = db.Column(db.Integer, db.ForeignKey('employe.id'), nullable=False)
    date_pointage = db.Column(db.Date, nullable=False, default=date.today)
    
    # Les 4 moments de badgeage
    arrivee_matin = db.Column(db.DateTime)
    depart_midi = db.Column(db.DateTime)
    arrivee_apres_midi = db.Column(db.DateTime)
    depart_soir = db.Column(db.DateTime)
    
    # Informations supplémentaires
    heures_travaillees = db.Column(db.Float, default=0)
    retard_matin = db.Column(db.Boolean, default=False)
    retard_apres_midi = db.Column(db.Boolean, default=False)
    absence = db.Column(db.Boolean, default=False)
    commentaire = db.Column(db.Text)
    
    # Contrainte d'unicité : un seul enregistrement par employé par jour
    __table_args__ = (db.UniqueConstraint('employe_id', 'date_pointage'),)

@login_manager.user_loader
def load_user(user_id):
    return Admin.query.get(int(user_id))

# Fonction pour calculer les heures travaillées
def calculer_heures(pointage):
    heures_total = 0
    
    # Période du matin
    if pointage.arrivee_matin and pointage.depart_midi:
        delta_matin = pointage.depart_midi - pointage.arrivee_matin
        heures_total += delta_matin.total_seconds() / 3600
    
    # Période de l'après-midi
    if pointage.arrivee_apres_midi and pointage.depart_soir:
        delta_apres_midi = pointage.depart_soir - pointage.arrivee_apres_midi
        heures_total += delta_apres_midi.total_seconds() / 3600
    
    return round(heures_total, 2)

# Routes principales
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/badge', methods=['GET', 'POST'])
def badge():
    if request.method == 'POST':
        matricule = request.form.get('matricule')
        type_badge = request.form.get('type_badge')
        
        employe = Employe.query.filter_by(matricule=matricule, actif=True).first()
        
        if not employe:
            flash('Matricule invalide ou employé inactif', 'error')
            return redirect(url_for('badge'))
        
        # Obtenir ou créer le pointage du jour
        aujourd_hui = date.today()
        pointage = Pointage.query.filter_by(
            employe_id=employe.id,
            date_pointage=aujourd_hui
        ).first()
        
        if not pointage:
            pointage = Pointage(
                employe_id=employe.id,
                date_pointage=aujourd_hui
            )
            db.session.add(pointage)
        
        # Enregistrer l'heure actuelle
        maintenant = datetime.now()
        
        # Logique de badgeage selon le type
        if type_badge == 'arrivee_matin':
            if pointage.arrivee_matin:
                flash('Vous avez déjà badgé votre arrivée ce matin', 'warning')
            else:
                pointage.arrivee_matin = maintenant
                # Vérifier le retard (après 9h00)
                heure_limite = datetime.strptime(app.config.get('HEURE_ARRIVEE_MAX', '09:00'), '%H:%M').time()
                if maintenant.time() > heure_limite:
                    pointage.retard_matin = True
                flash(f'Arrivée enregistrée à {maintenant.strftime("%H:%M")}', 'success')
        
        elif type_badge == 'depart_midi':
            if not pointage.arrivee_matin:
                flash('Vous devez d\'abord badger votre arrivée', 'error')
            elif pointage.depart_midi:
                flash('Vous avez déjà badgé votre départ midi', 'warning')
            else:
                pointage.depart_midi = maintenant
                flash(f'Départ midi enregistré à {maintenant.strftime("%H:%M")}', 'success')
        
        elif type_badge == 'arrivee_apres_midi':
            if pointage.arrivee_apres_midi:
                flash('Vous avez déjà badgé votre retour de pause', 'warning')
            else:
                pointage.arrivee_apres_midi = maintenant
                # Vérifier le retard (après 14h00)
                heure_limite = datetime.strptime(app.config.get('HEURE_RETOUR_MAX', '14:00'), '%H:%M').time()
                if maintenant.time() > heure_limite:
                    pointage.retard_apres_midi = True
                flash(f'Retour de pause enregistré à {maintenant.strftime("%H:%M")}', 'success')
        
        elif type_badge == 'depart_soir':
            if not pointage.arrivee_matin and not pointage.arrivee_apres_midi:
                flash('Vous devez d\'abord badger une arrivée', 'error')
            elif pointage.depart_soir:
                flash('Vous avez déjà badgé votre départ', 'warning')
            else:
                pointage.depart_soir = maintenant
                # Calculer les heures travaillées
                pointage.heures_travaillees = calculer_heures(pointage)
                flash(f'Départ enregistré à {maintenant.strftime("%H:%M")}. Total: {pointage.heures_travaillees}h', 'success')
        
        db.session.commit()
        return redirect(url_for('badge'))
    
    return render_template('badge.html')

# Routes espace employé
@app.route('/employe')
def employe_home():
    return render_template('employe_home.html')

@app.route('/employe/login', methods=['GET', 'POST'])
def employe_login():
    if request.method == 'POST':
        matricule = request.form.get('matricule')
        employe = Employe.query.filter_by(matricule=matricule, actif=True).first()
        
        if employe:
            session['employe_id'] = employe.id
            return redirect(url_for('employe_dashboard'))
        else:
            flash('Matricule invalide', 'error')
    
    return render_template('employe_login.html')

@app.route('/employe/dashboard')
def employe_dashboard():
    if 'employe_id' not in session:
        return redirect(url_for('employe_login'))
    
    employe = Employe.query.get(session['employe_id'])
    if not employe:
        session.pop('employe_id', None)
        return redirect(url_for('employe_login'))
    
    # Statistiques du mois en cours
    debut_mois = date.today().replace(day=1)
    pointages = Pointage.query.filter(
        Pointage.employe_id == employe.id,
        Pointage.date_pointage >= debut_mois
    ).order_by(Pointage.date_pointage.desc()).all()
    
    stats = {
        'heures_mois': sum(p.heures_travaillees for p in pointages),
        'retards': sum(1 for p in pointages if p.retard_matin or p.retard_apres_midi),
        'jours_travailles': len(pointages)
    }
    
    return render_template('employe_dashboard.html', 
                         employe=employe, 
                         pointages=pointages[:10],  # Derniers 10 pointages
                         stats=stats)

@app.route('/employe/logout')
def employe_logout():
    session.pop('employe_id', None)
    return redirect(url_for('index'))

# Routes administration
@app.route('/admin')
@login_required
def admin_dashboard():
    # Statistiques générales
    nb_employes = Employe.query.filter_by(actif=True).count()
    
    # Pointages du jour
    aujourd_hui = date.today()
    pointages_jour = db.session.query(Pointage, Employe).join(Employe).filter(
        Pointage.date_pointage == aujourd_hui
    ).all()
    
    # Calcul des présents/absents
    employes_actifs = Employe.query.filter_by(actif=True).all()
    employes_badges = [p.employe_id for p, e in pointages_jour]
    absents = [e for e in employes_actifs if e.id not in employes_badges]
    
    return render_template('admin_dashboard.html',
                         nb_employes=nb_employes,
                         pointages_jour=pointages_jour,
                         absents=absents,
                         aujourd_hui=aujourd_hui)

@app.route('/admin/employes')
@login_required
def liste_employes():
    employes = Employe.query.order_by(Employe.nom).all()
    return render_template('liste_employes.html', employes=employes)

@app.route('/admin/employe/ajouter', methods=['GET', 'POST'])
@login_required
def ajouter_employe():
    if request.method == 'POST':
        employe = Employe(
            matricule=request.form.get('matricule'),
            nom=request.form.get('nom'),
            prenom=request.form.get('prenom'),
            departement=request.form.get('departement'),
            email=request.form.get('email'),
            telephone=request.form.get('telephone')
        )
        
        try:
            db.session.add(employe)
            db.session.commit()
            flash('Employé ajouté avec succès', 'success')
            return redirect(url_for('liste_employes'))
        except:
            db.session.rollback()
            flash('Erreur: Le matricule existe déjà', 'error')
    
    return render_template('ajouter_employe.html')

@app.route('/admin/employe/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
def modifier_employe(id):
    employe = Employe.query.get_or_404(id)
    
    if request.method == 'POST':
        employe.matricule = request.form.get('matricule')
        employe.nom = request.form.get('nom')
        employe.prenom = request.form.get('prenom')
        employe.departement = request.form.get('departement')
        employe.email = request.form.get('email')
        employe.telephone = request.form.get('telephone')
        employe.actif = request.form.get('actif') == 'on'
        
        try:
            db.session.commit()
            flash('Employé modifié avec succès', 'success')
            return redirect(url_for('liste_employes'))
        except:
            db.session.rollback()
            flash('Erreur lors de la modification', 'error')
    
    return render_template('modifier_employe.html', employe=employe)

@app.route('/admin/employe/<int:id>/supprimer')
@login_required
def supprimer_employe(id):
    employe = Employe.query.get_or_404(id)
    
    try:
        db.session.delete(employe)
        db.session.commit()
        flash('Employé supprimé avec succès', 'success')
    except:
        db.session.rollback()
        flash('Erreur lors de la suppression', 'error')
    
    return redirect(url_for('liste_employes'))

@app.route('/admin/rapport/<type_rapport>')
@login_required
def generer_rapport(type_rapport):
    if type_rapport == 'jour':
        date_str = request.args.get('date', date.today().isoformat())
        date_rapport = datetime.fromisoformat(date_str).date()
        
        pointages = db.session.query(Pointage, Employe).join(Employe).filter(
            Pointage.date_pointage == date_rapport
        ).all()
        
        # Créer le fichier Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Pointages"
        
        # En-têtes
        headers = ['Matricule', 'Nom', 'Prénom', 'Département', 'Arrivée Matin', 
                   'Départ Midi', 'Arrivée Après-midi', 'Départ Soir', 
                   'Heures Travaillées', 'Retard Matin', 'Retard Après-midi']
        
        # Style pour les en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajouter les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Ajouter les données
        row_num = 2
        for p, e in pointages:
            ws.cell(row=row_num, column=1, value=e.matricule)
            ws.cell(row=row_num, column=2, value=e.nom)
            ws.cell(row=row_num, column=3, value=e.prenom)
            ws.cell(row=row_num, column=4, value=e.departement or '-')
            ws.cell(row=row_num, column=5, value=p.arrivee_matin.strftime('%H:%M') if p.arrivee_matin else 'Non badgé')
            ws.cell(row=row_num, column=6, value=p.depart_midi.strftime('%H:%M') if p.depart_midi else 'Non badgé')
            ws.cell(row=row_num, column=7, value=p.arrivee_apres_midi.strftime('%H:%M') if p.arrivee_apres_midi else 'Non badgé')
            ws.cell(row=row_num, column=8, value=p.depart_soir.strftime('%H:%M') if p.depart_soir else 'Non badgé')
            ws.cell(row=row_num, column=9, value=p.heures_travaillees)
            ws.cell(row=row_num, column=10, value='Oui' if p.retard_matin else 'Non')
            ws.cell(row=row_num, column=11, value='Oui' if p.retard_apres_midi else 'Non')
            row_num += 1
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder le fichier
        filename = f'rapport_pointage_{date_rapport.isoformat()}.xlsx'
        filepath = os.path.join('static', filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True)
    
    elif type_rapport == 'mensuel':
        mois = int(request.args.get('mois', date.today().month))
        annee = int(request.args.get('annee', date.today().year))
        
        # Créer le rapport mensuel
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport Mensuel"
        
        # En-têtes
        headers = ['Matricule', 'Nom', 'Prénom', 'Jours Travaillés', 
                   'Total Heures', 'Retards', 'Absences']
        
        # Style pour les en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Récupérer les données du mois
        employes = Employe.query.filter_by(actif=True).all()
        row_num = 2
        
        for employe in employes:
            pointages = Pointage.query.filter(
                Pointage.employe_id == employe.id,
                db.extract('month', Pointage.date_pointage) == mois,
                db.extract('year', Pointage.date_pointage) == annee
            ).all()
            
            jours_travailles = len(pointages)
            total_heures = sum(p.heures_travaillees for p in pointages)
            retards = sum(1 for p in pointages if p.retard_matin or p.retard_apres_midi)
            absences = sum(1 for p in pointages if p.absence)
            
            ws.cell(row=row_num, column=1, value=employe.matricule)
            ws.cell(row=row_num, column=2, value=employe.nom)
            ws.cell(row=row_num, column=3, value=employe.prenom)
            ws.cell(row=row_num, column=4, value=jours_travailles)
            ws.cell(row=row_num, column=5, value=round(total_heures, 2))
            ws.cell(row=row_num, column=6, value=retards)
            ws.cell(row=row_num, column=7, value=absences)
            row_num += 1
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder
        filename = f'rapport_mensuel_{annee}_{mois:02d}.xlsx'
        filepath = os.path.join('static', filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True)
    
    return redirect(url_for('admin_dashboard'))

@app.route('/login', methods=['GET', 'POST'])
@app.route('/admin-globibat', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('admin_dashboard'))
        
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        admin = Admin.query.filter_by(username=username).first()
        
        if admin and admin.check_password(password):
            login_user(admin, remember=True)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('admin_dashboard'))
        else:
            flash('Identifiants invalides', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

# API pour les statistiques
@app.route('/api/stats/dashboard')
@login_required
def api_stats_dashboard():
    aujourd_hui = date.today()
    
    # Nombre d'employés actifs
    employes_actifs = Employe.query.filter_by(actif=True).count()
    
    # Présents aujourd'hui
    presents_jour = db.session.query(Pointage).join(Employe).filter(
        Pointage.date_pointage == aujourd_hui,
        Employe.actif == True
    ).count()
    
    # Retards aujourd'hui
    retards_jour = db.session.query(Pointage).join(Employe).filter(
        Pointage.date_pointage == aujourd_hui,
        Employe.actif == True,
        db.or_(Pointage.retard_matin == True, Pointage.retard_apres_midi == True)
    ).count()
    
    # Heures moyennes (sur les 7 derniers jours)
    debut_semaine = aujourd_hui - timedelta(days=7)
    pointages_semaine = Pointage.query.filter(
        Pointage.date_pointage >= debut_semaine,
        Pointage.heures_travaillees > 0
    ).all()
    
    heures_moyennes = 0
    if pointages_semaine:
        heures_moyennes = sum(p.heures_travaillees for p in pointages_semaine) / len(pointages_semaine)
    
    return jsonify({
        'employes_actifs': employes_actifs,
        'presents_jour': presents_jour,
        'retards_jour': retards_jour,
        'heures_moyennes': round(heures_moyennes, 2)
    })

# Initialisation de la base de données
def init_db():
    with app.app_context():
        db.create_all()

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000) 